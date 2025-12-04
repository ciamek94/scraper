#!/usr/bin/env python3
# main.py
"""
Scraper OLX -> Excel + OneDrive + Telegram notifications (accepted/rejected).

Requirements:
  pip install requests beautifulsoup4 pandas openpyxl python-dotenv

Environment variables (set as GitHub Secrets):
  ONEDRIVE_CLIENT_ID
  ONEDRIVE_REFRESH_TOKEN
  TELEGRAM_BOT_TOKEN
  TELEGRAM_CHAT_ID

Optional:
  ONEDRIVE_UPLOAD_FOLDER (e.g. "olx") - folder name in OneDrive root where files will be stored
"""

import os
import time
import random
import json
import requests
import pandas as pd
import openpyxl
import re
import sys
import tempfile
import shutil
from bs4 import BeautifulSoup
from collections import defaultdict
from dotenv import load_dotenv

load_dotenv()

# --- Config (editable) ---
# List of searches: each search is a dict with 'name', 'urls' and filters
SEARCHES = [
    {
        "name": "falownik",
        "urls": [
            "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:to%5D=200",
            "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:from%5D=201&search%5Bfilter_float_price:to%5D=300",
            "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:from%5D=301&search%5Bfilter_float_price:to%5D=500",
            "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:from%5D=501&search%5Bfilter_float_price:to%5D=700",
            "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:from%5D=701&search%5Bfilter_float_price:to%5D=1000",
            "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:from%5D=1001&search%5Bfilter_float_price:to%5D=1400",
            # "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:from%5D=1401&search%5Bfilter_float_price:to%5D=2000",
            # "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:from%5D=2001&search%5Bfilter_float_price:to%5D=2500",
            # "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:from%5D=2501&search%5Bfilter_float_price:to%5D=3000",
            # "https://www.olx.pl/oferty/q-falownik/?search%5Bfilter_float_price:from%5D=3001&search%5Bfilter_float_price:to%5D=4000"

        ],
        "forbidden_words": [
            "fotowoltaiczny", "fotowoltaika", "fotowoltaiki", "fotowoltaicznej", "pv", "fotowoltaiczne", "fotowoltaiczna",
            "solar", "solarny", "magazyn energii", "mikroinwerter", "wifi",
            "off-grid", "on-grid", "off grid", "on grid", "offgrid", "ongrid",
            "hybrydowy", "hybrydowa", "solaredge", "deye", "growatt", "huawei",
            "sofar", "sma", "fox", "foxess", "fronius", "mppt", "easun",
            "sinuspro", "anern", "jebao", "godwe", "goodwe", "afore", "solis",
            "solax", "akwarium", "samochód", "toyota", "kia", "tunze", "opel",
            "audi", "volkswagen", "nissan", "victron", "solplanet", "sunny", "boy", "sunny boy", 
            "growat", "solax", "hypontech", "kempingowy", "suszarka", "pralka", "pompa ciepła",
            "anenji", "mercedes", "prius", "betoniarka", "słoneczne", "słoneczny", "volt polska",
            "cyrkulator", "ups", "akwariowa", "frezarka", "tokarka", "daye", "hoymiles", "dokio",
            "mieszkanie", "victron", "peleciarka", "samochodowa", "peugeot", "renault", "wavemaker",
            "bmw", "suzuki", "kodak", "kostal", "fox", "suszarke", "pralke", "ibo", "rolmasaż",
            "hypnotech", "greencell", "green cell", "masażer", "rubik", "lexus", "motech", "ford",
            "blaupunkt", "rollmasaż", "volvo", "still", "kamper", "bank energii", "zoe", "eclipse cross",
            "turystyczny", "hyundai", "suszarki", "pralki", "bosch", "chevrolet", "outlander", "dewalt", "makita",
            "milwuakee", "lodówka", "lodówki", "agregat", "spawarka", "spawarki", "refusol", "wilo", "glebogryzarka",
            "rower", "hulajnoga", "hulajnogi", "zoe", "kangoo", "grundfoss", "grundfos", "mazda", "tesla", "pulsor", "dacia",
            "walcarka", "do włosów", "płyta indukcyjna", "optymalizator", "optymalizatora", "turbiny wiatrowej", "turbina wiatrowa",
            "powerstocc", "milwaukee", "akwarystyczny", "akumulator", "belkin", "samochodowy", "chrysler", "aqua", "jvp", "SSR",
            "skoda", "mostek trapezowy", "remington", "aquael", "yaris", "willo", "wkład kominowy", "einhell", "simet", "honda",
            "osram", "lokówka", "prostownica", "lunchbox", "hydroforowy", "falownica", "dell", "kotła", "dzban", "sukienka",
            "odstraszacz", "klimatyzatora", "hybryda", "gyre", "konwerter", "sun lite", "aurora", "mokka", "sunwind", "rav 4",
            "vw id3", "vw id.3", "stecagrid", "micovert", "omnigena", "podgrzewacz", "Jaguar", "mikrofala", "tosima", "maszynka do mięsa",
            "citroen", "jeep", "linde", "steca", "mastervolt", "wózek widłowy", "wózka widłowego", "fluke", "maszynka do mielenia", "porsche",
            "selfa", "pompy ciepla", "maszynka alfa","kuchenka", "mikrofala", "mikrofalówka", "evershine"

            
        ],
        "required_words": [],  # if empty -> no requirement, otherwise at least one must appear
        "max_price": None,     # number or None
        "min_price": None
    },
    # {
    #     "name": "sprezarka",
    #     "urls": ["https://www.olx.pl/oferty/q-spre%C5%BCarka-%C5%9Srubowa/?search%5Bfilter_float_price:to%5D=6000",],
    #     "forbidden_words": ["wynajem,"],
    #     "required_words": [],
    #     "max_price": None,
    #     "min_price": None
    # }
]

WORKDIR = "output"
os.makedirs(WORKDIR, exist_ok=True)

CLIENT_ID = os.environ.get("ONEDRIVE_CLIENT_ID")
REFRESH_TOKEN = os.environ.get("ONEDRIVE_REFRESH_TOKEN")
TOKEN_URL = 'https://login.microsoftonline.com/consumers/oauth2/v2.0/token'
ONEDRIVE_UPLOAD_FOLDER = os.environ.get("ONEDRIVE_UPLOAD_FOLDER", "olx")

MAX_PAGES = 30
MAX_EMPTY_PAGES = 2

TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID")

# Missing-count behaviour: number of consecutive runs where an ad was NOT found.
# After exceeding KEEP_MISSING runs the ad will be removed.
MISSING_THRESHOLD = int(os.environ.get("KEEP_MISSING", "10"))
# OneDrive paths
EXCEL_ACCEPTED_ONEDRIVE = f"{ONEDRIVE_UPLOAD_FOLDER}/accepted.xlsx"
EXCEL_REJECTED_ONEDRIVE = f"{ONEDRIVE_UPLOAD_FOLDER}/rejected.xlsx"
JSON_ACCEPTED_ONEDRIVE = f"{ONEDRIVE_UPLOAD_FOLDER}/accepted.json"
JSON_REJECTED_ONEDRIVE = f"{ONEDRIVE_UPLOAD_FOLDER}/rejected.json"
STATE_ONEDRIVE_PATH = f"{ONEDRIVE_UPLOAD_FOLDER}/state.json"

# Local paths
EXCEL_ACCEPTED_LOCAL = os.path.join(WORKDIR, "accepted.xlsx")
EXCEL_REJECTED_LOCAL = os.path.join(WORKDIR, "rejected.xlsx")
JSON_ACCEPTED_LOCAL = os.path.join(WORKDIR, "accepted.json")
JSON_REJECTED_LOCAL = os.path.join(WORKDIR, "rejected.json")
STATE_LOCAL = os.path.join(WORKDIR, "state.json")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; olx-scraper/1.0)",
    "Accept-Language": "pl-PL,pl;q=0.9"
}

def get_with_retry(url, headers=HEADERS, retries=4, backoff=2.0):
    for i in range(retries):
        try:
            r = requests.get(url, headers=headers, timeout=15)
            # debug logging: status and short snippet to detect captcha/block
            if r.status_code != 200:
                print(f"⚠️ HTTP {r.status_code} for {url} (attempt {i+1}/{retries})")
                body = (r.text or "")[:300].replace("\n", " ")
                print("  snippet:", body)
            if r.status_code == 200:
                return r
        except Exception as e:
            print(f"⚠️ Request error for {url}: {e} (attempt {i+1}/{retries})")
        time.sleep(backoff * (1 + random.random()))
    print(f"❌ Giving up fetching {url} after {retries} attempts")
    return None

def update_env_refresh_token(new_token, filename_candidates=None):
    """Atomically update ONEDRIVE_REFRESH_TOKEN in .env with backup; do NOT print token."""
    try:
        if filename_candidates is None:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            filename_candidates = [
                os.path.join(base_dir, ".env"),
                os.path.join(os.getcwd(), ".env")
            ]
        env_path = next((p for p in filename_candidates if os.path.exists(p)), filename_candidates[0])
        # ensure file exists (create with secure perms if missing)
        if not os.path.exists(env_path):
            dirn = os.path.dirname(env_path) or "."
            os.makedirs(dirn, exist_ok=True)
            with open(env_path, "w", encoding="utf-8") as _:
                pass
            try:
                os.chmod(env_path, 0o600)
            except Exception:
                pass

        # read existing
        with open(env_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        # detect current value
        cur = None
        for ln in lines:
            if ln.strip().startswith("ONEDRIVE_REFRESH_TOKEN="):
                cur = ln.strip().split("=", 1)[1]
                break
        if cur == new_token:
            print("ℹ️ ONEDRIVE_REFRESH_TOKEN unchanged.")
            os.environ["ONEDRIVE_REFRESH_TOKEN"] = new_token
            globals()["REFRESH_TOKEN"] = new_token
            return True

        # backup current .env (keep single .env.bak to avoid filling disk)
        if os.environ.get("KEEP_ENV_BACKUP", "1") == "1":
            bak = env_path + ".bak"
            try:
                shutil.copy2(env_path, bak)  # overwrite single backup (no timestamp)
            except Exception as e:
                print("⚠️ Failed to create .env backup:", e)
        else:
            bak = None

        # build new content (replace or append)
        found = False
        out_lines = []
        for ln in lines:
            if ln.strip().startswith("ONEDRIVE_REFRESH_TOKEN="):
                out_lines.append(f"ONEDRIVE_REFRESH_TOKEN={new_token}\n")
                found = True
            else:
                out_lines.append(ln)
        if not found:
            out_lines.append(f"\n# updated by scraper\nONEDRIVE_REFRESH_TOKEN={new_token}\n")

        # atomic write
        dirn = os.path.dirname(env_path) or "."
        fd, tmp = tempfile.mkstemp(dir=dirn, prefix=".tmp_env_")
        os.close(fd)
        with open(tmp, "w", encoding="utf-8") as f:
            f.writelines(out_lines)
        os.replace(tmp, env_path)
        try:
            os.chmod(env_path, 0o600)
        except Exception:
            pass

        # update runtime env
        os.environ["ONEDRIVE_REFRESH_TOKEN"] = new_token
        globals()["REFRESH_TOKEN"] = new_token
        print("ℹ️ ONEDRIVE_REFRESH_TOKEN updated in .env (backup saved).")
        return True
    except Exception as e:
        print("⚠️ Failed to update .env with new refresh_token:", e)
        return False

def authenticate_onedrive():
    if not CLIENT_ID or not REFRESH_TOKEN:
        print("⚠️ OneDrive credentials not set.")
        return None
    data = {
        'client_id': CLIENT_ID,
        'refresh_token': REFRESH_TOKEN,
        'grant_type': 'refresh_token',
        'scope': 'offline_access Files.ReadWrite.All openid profile'
    }
    try:
        r = requests.post(TOKEN_URL, data=data, timeout=20)
        r.raise_for_status()
        j = r.json()
        at = j.get("access_token", "")
        if not at:
            print("❌ OneDrive auth returned no access_token. Response:", j)
            return None
        # verify token works by calling a small Graph endpoint
        headers = {"Authorization": f"Bearer {at}"}
        try:
            test = requests.get("https://graph.microsoft.com/v1.0/me/drive", headers=headers, timeout=10)
            if test.status_code not in (200, 201):
                print("❌ Graph API rejected token:", test.status_code, test.text)
                print("Token endpoint response:", j)
                return None
        except Exception as e:
            print("❌ Graph API test request failed:", e)
            return None
        print("✅ OneDrive auth successful. Access token accepted by Graph.")
        # if server returned a new refresh_token -> persist it to .env
        new_rt = j.get("refresh_token")
        if new_rt:
            # do not print the token value
            ok = update_env_refresh_token(new_rt)
            if not ok:
                print("⚠️ Could not persist new refresh_token to .env (check permissions).")
        return j
    except requests.exceptions.RequestException as e:
        print("❌ OneDrive auth failed:", e, r.text if 'r' in locals() else "")
        return None

def upload_to_onedrive_localpath(local_path, onedrive_path, token):
    if token is None:
        print("⚠️ No OneDrive token, skipping upload:", onedrive_path)
        return False
    access_token = token['access_token']
    upload_url = f'https://graph.microsoft.com/v1.0/me/drive/root:/{onedrive_path}:/content'
    headers = {'Authorization': f"Bearer {access_token}"}
    with open(local_path, "rb") as f:
        data = f.read()
    r = requests.put(upload_url, headers=headers, data=data, timeout=60)
    if r.status_code in (200, 201):
        print("✅ Uploaded to OneDrive:", onedrive_path)
        return True
    else:
        print("❌ Upload failed:", r.status_code, r.text)
        return False

def download_from_onedrive(onedrive_path, local_path, token):
    if token is None:
        print("⚠️ No OneDrive token, cannot download", onedrive_path)
        return False
    access_token = token['access_token']
    url = f'https://graph.microsoft.com/v1.0/me/drive/root:/{onedrive_path}:/content'
    headers = {'Authorization': f'Bearer {access_token}'}
    r = requests.get(url, headers=headers, timeout=60)
    if r.status_code == 200:
        with open(local_path, "wb") as f:
            f.write(r.content)
        print("✅ Downloaded from OneDrive:", onedrive_path)
        return True
    else:
        print("ℹ️ File not found on OneDrive (or download failed):", onedrive_path, r.status_code)
        return False

def send_telegram_notification(title, price, link, photo_url=None):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("⚠️ Telegram not configured — skipping notification.")
        return False
    base = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"
    caption = f"<b>{title}</b>\n{price}\n{link}"
    if photo_url:
        try:
            r = requests.post(f"{base}/sendPhoto", data={
                "chat_id": TELEGRAM_CHAT_ID,
                "photo": photo_url,
                "caption": caption,
                "parse_mode": "HTML",
                "disable_web_page_preview": False
            }, timeout=15)
            if r.status_code == 200:
                return True
        except Exception:
            pass
    try:
        r = requests.post(f"{base}/sendMessage", data={
            "chat_id": TELEGRAM_CHAT_ID,
            "text": caption,
            "parse_mode": "HTML",
            "disable_web_page_preview": False
        }, timeout=10)
        return r.status_code == 200
    except Exception:
        return False

def clean_price(price_str):
    if not price_str:
        return ""
    return price_str.replace("\n", " ").strip()

def normalize_price(price_str):
    """Return integer price if digits found, otherwise None."""
    if not price_str:
        return None
    # find first group of digits (handles spaces like "1 200")
    m = re.search(r"(\d[\d\s]*)", str(price_str))
    if not m:
        return None
    try:
        return int(m.group(1).replace(" ", ""))
    except Exception:
        return None

def is_negotiable(price_str):
    """Detect Polish 'do negocjacji' (case-insensitive, tolerant to spacing/typos)."""
    if not price_str:
        return False
    return bool(re.search(r"do\s*negocj", price_str, re.IGNORECASE))

def parse_search_page(html):
    soup = BeautifulSoup(html, "html.parser")
    # Collect all possible ad containers (l-card, ad-card-title, premium-ad-card, any with "card" in data-cy)
    cards = soup.find_all("div", {"data-cy": re.compile(r"card")})
    results = []
    for card in cards:
        # link + title
        link_elem = card.find("a", href=True)
        title_elem = card.find("h4")
        title = title_elem.get_text(strip=True) if title_elem else ""
        link = link_elem["href"] if link_elem else ""
        if link and not link.startswith("http"):
            link = "https://www.olx.pl" + link

        # price
        price_elem = card.find("p", {"data-testid": "ad-price"})
        price = clean_price(price_elem.get_text(strip=True)) if price_elem else ""

        # location and date
        loc_date_elem = card.find("p", {"data-testid": "location-date"})
        loc_date = loc_date_elem.get_text(" ", strip=True) if loc_date_elem else ""

        results.append({
            "title": title,
            "link": link,
            "price": price,
            "loc_date": loc_date
        })

    # Get the total number of ads from OLX (e.g. "We found 797 ads")
    count_elem = soup.find("span", {"data-testid": "total-count"})
    if count_elem:
        match = re.search(r"(\d[\d\s]*)", count_elem.get_text())
        if match:
            total_count = int(match.group(1).replace(" ", ""))
            print(f"ℹ️ OLX reports {total_count} ads in the entire search, scraper found {len(results)} on this page.")
        else:
            print(f"ℹ️ Scraper found {len(results)} ads on this page (could not read OLX count).")
    else:
        print(f"ℹ️ Scraper found {len(results)} ads on this page (OLX counter not found).")

    return results

def parse_listing_page(html):
    soup = BeautifulSoup(html, "html.parser")
    desc_elem = soup.find("div", {"data-cy": "ad_description"})
    if not desc_elem:
        desc_elem = soup.find("div", {"class": lambda x: x and "description" in x})
    description = desc_elem.get_text(" ", strip=True) if desc_elem else ""
    image_url = None
    meta_img = soup.find("meta", property="og:image")
    if meta_img and meta_img.has_attr("content"):
        image_url = meta_img["content"]
    else:
        img = soup.find("img", {"class": lambda x: x and ("swiper" in x or "image" in x or "gallery" in x)})
        if img and img.has_attr("src"):
            image_url = img["src"]
        else:
            gallery_img = soup.select_one("div.photos img")
            if gallery_img and gallery_img.has_attr("src"):
                image_url = gallery_img["src"]
    return description, image_url

def normalize_text(text):
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def passes_filters(item, search_conf):
    text = normalize_text(item.get("title","") + " " + item.get("description",""))
    for bad in search_conf.get("forbidden_words", []):
        if normalize_text(bad) in text:
            return False
    reqs = search_conf.get("required_words", [])
    if reqs and not any(normalize_text(r) in text for r in reqs):
        return False
    price = item.get("price","")
    if price:
        digits = "".join(ch for ch in price if ch.isdigit())
        if digits:
            try:
                pnum = int(digits)
                maxp = search_conf.get("max_price")
                minp = search_conf.get("min_price")
                if maxp is not None and pnum > maxp:
                    return False
                if minp is not None and pnum < minp:
                    return False
            except Exception:
                pass
    return True

# Load/Save helpers for Excel/JSON
def load_json(path):
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_json(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_excel(path):
    if os.path.exists(path):
        try:
            return pd.read_excel(path)
        except Exception:
            return pd.DataFrame()
    return pd.DataFrame()

def save_excel(df, path):
    df.to_excel(path, index=False)
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
        wb.save(path)
    except Exception as e:
        print("⚠️ autosize failed:", e)

def normalize_link(url):
    """
    Normalize OLX listing URL for stable deduplication:
    - ensure scheme (default https)
    - normalize host (m.olx.pl -> www.olx.pl)
    - remove query string and fragment
    - strip trailing slash
    """
    if not url:
        return None
    try:
        from urllib.parse import urlparse
        p = urlparse(url)
    except Exception:
        return url.strip()
    scheme = p.scheme or "https"
    netloc = p.netloc.replace("m.olx.pl", "www.olx.pl")
    path = p.path or ""
    # remove trailing slash
    path = re.sub(r"/+$", "", path)
    if not path:
        path = "/"
    return f"{scheme}://{netloc}{path}"

def abort_with_notification(msg):
    print("❌ ABORT:", msg)
    try:
        # Try to send a notification via Telegram (if configured)
        send_telegram_notification("OLX scraper error", msg, "")
    except Exception:
        pass
    sys.exit(1)

def atomic_save_json(data, path):
    dirn = os.path.dirname(path) or "."
    fd, tmp = tempfile.mkstemp(dir=dirn, prefix=".tmp_")
    os.close(fd)
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)

def atomic_save_excel(df, path):
    # Save first to a temporary file, then replace the target file atomically
    dirn = os.path.dirname(path) or "."
    fd, tmp = tempfile.mkstemp(dir=dirn, suffix=".xlsx", prefix=".tmp_")
    os.close(fd)
    df.to_excel(tmp, index=False)
    try:
        wb = openpyxl.load_workbook(tmp)
        ws = wb.active
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
        wb.save(tmp)
    except Exception:
        pass
    os.replace(tmp, path)

def write_temp_json(data, target_path):
    dirn = os.path.dirname(target_path) or "."
    fd, tmp = tempfile.mkstemp(dir=dirn, prefix=".tmp_", suffix=".json")
    os.close(fd)
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return tmp

def write_temp_excel(df, target_path):
    dirn = os.path.dirname(target_path) or "."
    fd, tmp = tempfile.mkstemp(dir=dirn, prefix=".tmp_", suffix=".xlsx")
    os.close(fd)
    df.to_excel(tmp, index=False)
    try:
        wb = openpyxl.load_workbook(tmp)
        ws = wb.active
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2
        wb.save(tmp)
    except Exception:
        pass
    return tmp

def upload_temps_and_commit(token, tmp_map):
    """
    tmp_map: list of tuples (tmp_local_path, final_local_path, onedrive_path)
    Upload each tmp to OneDrive; if all succeed, replace final_local_path with tmp (os.replace).
    On any failure, delete tmp files and return False.
    """
    if token is None:
        print("⚠️ No OneDrive token, skipping upload/commit.")
        return False

    access_token = token.get("access_token")
    if not access_token:
        print("❌ Missing access_token, aborting upload.")
        return False

    # quick diagnostic: show token shape (no secret printed)
    dots = access_token.count('.')
    print(f"ℹ️ access_token length={len(access_token)} dots={dots}")

    headers = {'Authorization': f'Bearer {access_token}'}

    # verify token is accepted by Graph before uploading files
    try:
        test = requests.get("https://graph.microsoft.com/v1.0/me/drive", headers=headers, timeout=10)
        if test.status_code not in (200, 201):
            print("❌ Graph API rejected token before upload:", test.status_code, test.text)
            return False
    except Exception as e:
        print("❌ Graph API test request failed before upload:", e)
        return False

    uploaded = []
    try:
        for tmp_local, final_local, onedrive_path in tmp_map:
            upload_url = f'https://graph.microsoft.com/v1.0/me/drive/root:/{onedrive_path}:/content'
            with open(tmp_local, "rb") as f:
                data = f.read()
            r = requests.put(upload_url, headers=headers, data=data, timeout=60)
            if r.status_code not in (200, 201):
                print("❌ Upload failed:", r.status_code, r.text)
                raise RuntimeError(f"Upload failed for {onedrive_path}")
            print("✅ Uploaded to OneDrive:", onedrive_path)
            uploaded.append((tmp_local, final_local))
        # all uploads succeeded -> move tmp to final local paths atomically
        for tmp_local, final_local in uploaded:
            os.replace(tmp_local, final_local)
        return True
    except Exception as e:
        print("❌ Upload process error:", e)
        # cleanup tmp files
        for tmp_local, _, _ in tmp_map:
            try:
                if os.path.exists(tmp_local):
                    os.remove(tmp_local)
            except Exception:
                pass
        return False

# ---- Main run ----
def main():
    __version__ = "1.0.3"
    __version_date__ = "2025-12-04"
    print(f"main.py v{__version__} ({__version_date__})")

    print("🚀 OLX scraper starting")
    token = authenticate_onedrive() if (CLIENT_ID and REFRESH_TOKEN) else None

    # If we have OneDrive token, try to download remote files.
    # If any download fails AND local file doesn't exist -> abort (and notify).
    if token:
        downloads = [
            (STATE_ONEDRIVE_PATH, STATE_LOCAL),
            (EXCEL_ACCEPTED_ONEDRIVE, EXCEL_ACCEPTED_LOCAL),
            (EXCEL_REJECTED_ONEDRIVE, EXCEL_REJECTED_LOCAL),
            (JSON_ACCEPTED_ONEDRIVE, JSON_ACCEPTED_LOCAL),
            (JSON_REJECTED_ONEDRIVE, JSON_REJECTED_LOCAL),
        ]
        for remote, local in downloads:
            ok = download_from_onedrive(remote, local, token)
            if not ok and not os.path.exists(local):
                abort_with_notification(f"Failed to download required file from OneDrive: {remote} and local {local} missing. Aborting to avoid corrupting data.")
    else:
        print("⚠️ No OneDrive token — using local files if present.")

    # Load previous state if it exists. Keep state_raw = None when no previous state file.
    state_raw = None
    if os.path.exists(STATE_LOCAL):
        try:
            state_raw = load_json(STATE_LOCAL)
            if not isinstance(state_raw, dict) or not state_raw:
                state_raw = None
        except Exception:
            state_raw = None

    # use loaded state when available, otherwise initialize defaults (but mark as no previous run)
    state = state_raw if state_raw else {"seen": [], "last_prices": {}, "last_run": None}

    accepted_json = load_json(JSON_ACCEPTED_LOCAL)
    rejected_json = load_json(JSON_REJECTED_LOCAL)
    accepted_df = load_excel(EXCEL_ACCEPTED_LOCAL)
    rejected_df = load_excel(EXCEL_REJECTED_LOCAL)

    # Ensure MissingCount exists for existing entries
    for row in accepted_json:
        if "MissingCount" not in row:
            row["MissingCount"] = 0
    for row in rejected_json:
        if "MissingCount" not in row:
            row["MissingCount"] = 0

    # build maps using normalized links as keys (use stored NormLink when available)
    accepted_map = {(row.get('NormLink') or normalize_link(row.get('Link'))): row for row in accepted_json if row.get('Link')}
    rejected_map = {(row.get('NormLink') or normalize_link(row.get('Link'))): row for row in rejected_json if row.get('Link')}

    last_prices = state.get("last_prices", {})

    # current_links_found will store normalized links
    current_links_found = set()
    # seen_in_run avoids processing the same (normalized) link multiple times in one run
    seen_in_run = set()
    new_accepted = []
    new_rejected = []
    price_changed = []

    for search_conf in SEARCHES:
        name = search_conf["name"]
        urls = search_conf.get("urls", [search_conf.get("url")])
        for base_url in urls:
            if not base_url:
                continue
            print(f"🔎 Searching '{name}' at {base_url}")

            page = 1
            empty_pages = 0
            all_results = []  # Collect all ads for this search URL
            while page <= MAX_PAGES and empty_pages < MAX_EMPTY_PAGES:
                paged = base_url + (f"&page={page}" if "?" in base_url else f"?page={page}")
                print(" - Fetching", paged)
                r = get_with_retry(paged)
                if r is None:
                    empty_pages += 1
                    page += 1
                    time.sleep(random.uniform(1.5, 3.5))
                    continue

                results = parse_search_page(r.text)
                if not results:
                    empty_pages += 1
                    page += 1
                    time.sleep(random.uniform(1.0, 2.5))
                    continue

                all_results.extend(results)
                empty_pages = 0

                for res in results:
                    raw_link = res.get("link")
                    if not raw_link:
                        continue

                    norm_link = normalize_link(raw_link)
                    if not norm_link:
                        continue

                    # avoid duplicate processing within this run
                    if norm_link in seen_in_run:
                        continue
                    seen_in_run.add(norm_link)
                    current_links_found.add(norm_link)

                    # Check in accepted/rejected using normalized price comparison
                    price_raw = res.get("price")
                    price_num = normalize_price(price_raw)
                    negotiable = is_negotiable(price_raw)

                    acc_row = accepted_map.get(norm_link)
                    rej_row = rejected_map.get(norm_link)
                    acc_price_raw = acc_row.get("Price") if acc_row else None
                    rej_price_raw = rej_row.get("Price") if rej_row else None
                    acc_price_num = normalize_price(acc_price_raw)
                    rej_price_num = normalize_price(rej_price_raw)

                    def prices_equal(a_num, a_raw, b_num, b_raw):
                        # Prefer numeric comparison when both available, fallback to raw string compare
                        if a_num is not None and b_num is not None:
                            return a_num == b_num
                        return (a_raw or "").strip() == (b_raw or "").strip()

                    in_accepted = acc_row is not None and prices_equal(acc_price_num, acc_price_raw, price_num, price_raw)
                    in_rejected = rej_row is not None and prices_equal(rej_price_num, rej_price_raw, price_num, price_raw)
                    price_diff = acc_row is not None and not prices_equal(acc_price_num, acc_price_raw, price_num, price_raw)

                    if (acc_row or rej_row) and (not in_accepted and not in_rejected):
                        # existing record present but price differs -> debug info
                        stored = acc_price_raw or rej_price_raw
                        print(f"ℹ️ Existing record for {norm_link} found but price differs (stored: {stored} vs current: {price_raw}).")

                    if in_accepted or in_rejected:
                        # Skip fetching the listing page
                        continue

                    # Fetch listing page
                    lr = get_with_retry(raw_link)
                    if lr is None:
                        continue
                    description, image_url = parse_listing_page(lr.text)
                    res["description"] = description
                    res["image"] = image_url
                    res["search_name"] = name

                    if passes_filters(res, search_conf):
                        # Accepted
                        row = {
                            "Title": res.get("title",""),
                            "Price": price_raw,
                            "Negotiable": negotiable,
                            "Location/Date": res.get("loc_date",""),
                            "Description": res.get("description",""),
                            "Link": raw_link,
                            "NormLink": norm_link,
                            "Image": res.get("image"),
                            "SearchName": name,
                            "Notified": False,
                            "MissingCount": 0,
                            "Timestamp": int(time.time())
                        }
                        accepted_json.append(row)
                        accepted_map[norm_link] = row
                        new_accepted.append(row)
                        if price_diff:
                            row["Title"] += " ⚠️ Price changed"
                            price_changed.append(row)
                    else:
                        # Rejected
                        row = {
                            "Title": res.get("title",""),
                            "Price": price_raw,
                            "Negotiable": negotiable,
                            "Location/Date": res.get("loc_date",""),
                            "Description": description,
                            "Link": raw_link,
                            "NormLink": norm_link,
                            "Image": image_url,
                            "SearchName": name,
                            "MissingCount": 0,
                            "Timestamp": int(time.time())
                        }
                        rejected_json.append(row)
                        rejected_map[norm_link] = row
                        new_rejected.append(row)

                    # store numeric price when possible
                    last_prices[norm_link] = price_num if price_num is not None else (price_raw or "")

                    time.sleep(random.uniform(0.8, 1.8))
                page += 1
                time.sleep(random.uniform(1.5, 3.0))

            # --- SUMMARY FOR THIS SEARCH URL ---
            unique_links = set(normalize_link(ad["link"]) for ad in all_results if ad.get("link"))
            print(f"\n📊 Summary for '{name}' ({base_url}):")
            print(f"Scraper found {len(all_results)} ads (raw, all pages).")
            print(f"Scraper found {len([u for u in unique_links if u])} unique ads (across all pages).\n")

    # --- REMOVE/UPDATE ENTRIES NOT FOUND IN CURRENT RUN ---
    # Update MissingCount for entries not found in current run.
    def update_missing_counters(json_list, threshold=MISSING_THRESHOLD):
        kept = []
        removed = 0
        for row in json_list:
            link = normalize_link(row.get("Link"))
            if link in current_links_found:
                row["MissingCount"] = 0
                kept.append(row)
            else:
                row["MissingCount"] = int(row.get("MissingCount", 0)) + 1
                if row["MissingCount"] >= threshold:
                    removed += 1
                    # drop the row
                else:
                    kept.append(row)
        return kept, removed

    # Only update/remove if we had previous state (to avoid purging on first run)
    if state_raw:
        before_a = len(accepted_json)
        before_r = len(rejected_json)
        accepted_json, removed_a = update_missing_counters(accepted_json, MISSING_THRESHOLD)
        rejected_json, removed_r = update_missing_counters(rejected_json, MISSING_THRESHOLD)
        print(f"ℹ️ Removed {removed_a} accepted entries and {removed_r} rejected entries (MissingCount >= {MISSING_THRESHOLD}).")
    else:
        print("ℹ️ No previous state — skipping removal/update of MissingCount on first run.")

    # When adding new accepted entries, ensure MissingCount is initialised
    for row in new_accepted:
        if "MissingCount" not in row:
            row["MissingCount"] = 0

    # When adding new rejected entries, ensure MissingCount is initialised
    for row in new_rejected:
        if "MissingCount" not in row:
            row["MissingCount"] = 0

    # Save state locally only AFTER successful upload to OneDrive.
    state = {"seen": list(current_links_found), "last_prices": last_prices, "last_run": int(time.time())}

    # If we have OneDrive token -> prepare tmp files + upload, commit only on success.
    if token:
        # Prepare temp files for upload/commit
        tmp_state = write_temp_json(state, STATE_LOCAL)
        tmp_acc_json = write_temp_json(accepted_json, JSON_ACCEPTED_LOCAL)
        tmp_rej_json = write_temp_json(rejected_json, JSON_REJECTED_LOCAL)
        df_acc = pd.DataFrame(accepted_json)
        df_rej = pd.DataFrame(rejected_json)
        tmp_acc_xlsx = write_temp_excel(df_acc, EXCEL_ACCEPTED_LOCAL)
        tmp_rej_xlsx = write_temp_excel(df_rej, EXCEL_REJECTED_LOCAL)

        # Mapping: (tmp_local, final_local, onedrive_path)
        tmp_map = [
            (tmp_acc_xlsx, EXCEL_ACCEPTED_LOCAL, EXCEL_ACCEPTED_ONEDRIVE),
            (tmp_rej_xlsx, EXCEL_REJECTED_LOCAL, EXCEL_REJECTED_ONEDRIVE),
            (tmp_acc_json, JSON_ACCEPTED_LOCAL, JSON_ACCEPTED_ONEDRIVE),
            (tmp_rej_json, JSON_REJECTED_LOCAL, JSON_REJECTED_ONEDRIVE),
            (tmp_state, STATE_LOCAL, STATE_ONEDRIVE_PATH),
        ]

        # Refresh token right before upload (in case previous token expired during scraping)
        refreshed = authenticate_onedrive()
        if not refreshed:
            abort_with_notification("OneDrive auth failed before upload — aborting without modifying local files.")
        ok = upload_temps_and_commit(refreshed, tmp_map)
        if not ok:
            abort_with_notification("OneDrive upload failed — aborting without modifying local files.")
        # on success upload_temps_and_commit already replaced temps -> local files committed
    else:
        # No token -> commit locally immediately (atomic)
        df_acc = pd.DataFrame(accepted_json)
        df_rej = pd.DataFrame(rejected_json)
        atomic_save_json(state, STATE_LOCAL)
        atomic_save_json(accepted_json, JSON_ACCEPTED_LOCAL)
        atomic_save_json(rejected_json, JSON_REJECTED_LOCAL)
        atomic_save_excel(df_acc, EXCEL_ACCEPTED_LOCAL)
        atomic_save_excel(df_rej, EXCEL_REJECTED_LOCAL)

    # 🔄 Notifications
    to_notify = new_accepted + price_changed
    if to_notify:
        print(f"🔔 New accepted listings or price changes: {len(to_notify)} - sending notifications")
        for item in to_notify:
            ok = send_telegram_notification(
                title=item["Title"],
                price=item["Price"],
                link=item["Link"],
                photo_url=item.get("Image")
            )
            item["Notified"] = ok
            time.sleep(1.2 + random.random())
    else:
        print("ℹ️ No new accepted listings or price changes")

    print("✅ Done.")

if __name__ == "__main__":
    main()
